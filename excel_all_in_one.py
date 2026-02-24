# -*- coding: utf-8 -*-
"""
Excel All-in-One Tool v5.0
รวม 5 เครื่องมือ:
1. กรอกข้อมูล 3T          - กรอกรหัส+รายละเอียด ลง Column O-BI
2. กรอกข้อมูล Output      - บันทึกรายงานการโทร + รวมคอลัมน์ F-U -> F
3. ดึงข้อมูล 3T           - นับสายการโทร + เรียงลำดับ
4. ดึงข้อมูลจาก Output    - ค้นหารหัส + ดึงข้อมูล Column A,B,C,D,F
5. เปิดบิล               - กระดาษซับ / กระดาษรองรีด / อ่านบิลจากรูป
"""
import threading
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.toast import ToastNotification
from ttkbootstrap.scrolled import ScrolledText
from ttkbootstrap.dialogs import Messagebox
from tkinter import filedialog
import tkinter as tk
import openpyxl
import pandas as pd
import os
from PIL import Image, ImageTk, ImageEnhance

# ============================================================================
# CONSTANTS
# ============================================================================
SP = 12
FONT_FAMILY = "Segoe UI"
FONT_SIZE = 11
FONT_SIZE_TITLE = 15
ROW_START = 2
ROW_END = 553
BG_OPACITY = 0.15  # ความสว่างของ background (0=ดำสนิท, 1=สว่างเต็ม)


# ============================================================================
# UTILITIES
# ============================================================================
def center_window(win, w, h):
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    x = (sw - w) // 2
    y = (sh - h) // 2
    win.geometry(f"{w}x{h}+{x}+{y}")


def fit_window(win, min_w=600, min_h=400):
    """วัดขนาด content จริง แล้วปรับหน้าต่างให้พอดี (ไม่น้อยกว่า min_w/min_h)"""
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    req_w = min(max(win.winfo_reqwidth() + 24, min_w), int(sw * 0.95))
    req_h = min(max(win.winfo_reqheight() + 24, min_h), int(sh * 0.92))
    x = (sw - req_w) // 2
    y = (sh - req_h) // 2
    win.geometry(f"{req_w}x{req_h}+{x}+{y}")


def show_toast(title, message, duration=3000):
    toast = ToastNotification(
        title=title, message=message, duration=duration, bootstyle=SUCCESS
    )
    toast.show_toast()


def apply_table_style(tree):
    tree.tag_configure("oddrow", background="#1a1a1a")
    tree.tag_configure("evenrow", background="#262626")
    for i, item in enumerate(tree.get_children()):
        tree.item(item, tags=("evenrow",) if i % 2 == 0 else ("oddrow",))


# ============================================================================
# BACKGROUND IMAGE HELPER
# ============================================================================
_BG_ORIG = None  # cache รูปเพื่อไม่ต้องโหลดซ้ำทุกหน้าต่าง


def apply_bg(win):
    """ใส่รูป background โปร่งแสงให้กับหน้าต่างใดก็ได้"""
    global _BG_ORIG
    if _BG_ORIG is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        for name in ("bg.png", "bg.jpg", "bg.jpeg", "bg.webp"):
            p = os.path.join(script_dir, name)
            if os.path.exists(p):
                try:
                    _BG_ORIG = Image.open(p).convert("RGB")
                except Exception:
                    pass
                break
    if _BG_ORIG is None:
        return
    orig = _BG_ORIG
    lbl = tk.Label(win, bd=0, highlightthickness=0)
    lbl.place(x=0, y=0, relwidth=1, relheight=1)
    lbl.lower()

    def _update():
        w = win.winfo_width()
        h = win.winfo_height()
        if w < 10 or h < 10:
            return
        img = orig.copy()
        img_ratio = img.width / img.height
        win_ratio = w / h
        if img_ratio > win_ratio:
            new_h = h
            new_w = int(h * img_ratio)
        else:
            new_w = w
            new_h = int(w / img_ratio)
        img = img.resize((new_w, new_h), Image.LANCZOS)
        left = (new_w - w) // 2
        top = (new_h - h) // 2
        img = img.crop((left, top, left + w, top + h))
        img = ImageEnhance.Brightness(img).enhance(BG_OPACITY)
        photo = ImageTk.PhotoImage(img)
        lbl.configure(image=photo)
        lbl._photo = photo  # ป้องกัน garbage collection

    def _on_resize(event):
        if hasattr(lbl, '_job'):
            win.after_cancel(lbl._job)
        lbl._job = win.after(150, _update)

    win.after(50, _update)
    win.bind("<Configure>", _on_resize)


# ============================================================================
# TOOL 1: กรอกข้อมูล 3T
# ============================================================================
class Fill3TWindow:
    """กรอกรหัส + รายละเอียด ลงไฟล์ Excel (Column O-BI)"""

    def __init__(self, parent):
        self.win = ttk.Toplevel(parent)
        self.win.title("กรอกข้อมูล 3T - Excel Auto Mapper")
        self.file_path = None
        self._build_ui()
        fit_window(self.win, min_w=920, min_h=660)
        apply_bg(self.win)

    def _build_ui(self):
        # --- File selection ---
        file_frame = ttk.Frame(self.win, style="Card.TFrame", padding=SP)
        file_frame.pack(fill=X, padx=SP, pady=SP)

        ttk.Label(
            file_frame,
            text="เลือกไฟล์ Excel ต้นฉบับ:",
            font=(FONT_FAMILY, FONT_SIZE, "bold"),
        ).pack(side=LEFT, padx=(SP, 0))

        self.entry_path = ttk.Entry(file_frame, width=50)
        self.entry_path.pack(side=LEFT, padx=SP)

        ttk.Button(
            file_frame, text="Browse", command=self._browse_file, bootstyle=WARNING
        ).pack(side=LEFT)

        # --- Input area ---
        input_frame = ttk.Frame(self.win)
        input_frame.pack(fill=BOTH, expand=YES, padx=SP, pady=(0, SP))
        input_frame.columnconfigure(0, weight=1)
        input_frame.columnconfigure(1, weight=1)
        input_frame.rowconfigure(1, weight=1)

        ttk.Label(input_frame, text="กรอก 'รหัส' (Column A)").grid(
            row=0, column=0, pady=5
        )
        ttk.Label(input_frame, text="กรอก 'รายละเอียด' (ที่จะเติม)").grid(
            row=0, column=1, pady=5
        )

        self.txt_ids = ScrolledText(input_frame, height=20, autohide=True)
        self.txt_ids.grid(row=1, column=0, padx=(0, SP // 2), sticky="nsew")

        self.txt_details = ScrolledText(input_frame, height=20, autohide=True)
        self.txt_details.grid(row=1, column=1, padx=(SP // 2, 0), sticky="nsew")

        # --- Process button ---
        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(fill=X, padx=SP, pady=(0, SP))

        ttk.Button(
            btn_frame,
            text="เริ่มทำงาน (Start Process)",
            command=self._start_process,
            bootstyle=SUCCESS,
            width=40,
        ).pack(fill=X, ipady=5)

        self.lbl_status = ttk.Label(
            btn_frame, text="สถานะ: รอคำสั่ง...", foreground="gray"
        )
        self.lbl_status.pack(pady=5)

    def _browse_file(self):
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx;*.xlsm")]
        )
        if filename:
            self.file_path = filename
            self.entry_path.delete(0, "end")
            self.entry_path.insert(0, filename)
            self.lbl_status.configure(
                text="โหลดไฟล์เรียบร้อย พร้อมกรอกข้อมูล", foreground="cyan"
            )

    def _start_process(self):
        if not self.file_path:
            Messagebox.show_error("กรุณาเลือกไฟล์ Excel ก่อน", title="Error")
            return

        raw_ids = self.txt_ids.get("1.0", "end").strip().split("\n")
        raw_details = self.txt_details.get("1.0", "end").strip().split("\n")

        list_ids = [x.strip() for x in raw_ids if x.strip()]
        list_details = [x.strip() for x in raw_details if x.strip()]

        if len(list_ids) != len(list_details):
            Messagebox.show_warning(
                f"จำนวนบรรทัดไม่เท่ากัน!\n"
                f"รหัส: {len(list_ids)} รายการ\n"
                f"รายละเอียด: {len(list_details)} รายการ\n"
                f"โปรแกรมจะทำเฉพาะคู่ที่มีครบ",
                title="Warning",
            )

        pairs = list(zip(list_ids, list_details))
        if not pairs:
            Messagebox.show_info("ไม่มีข้อมูลให้ประมวลผล", title="Info")
            return

        try:
            self.lbl_status.configure(text="กำลังประมวลผล...", foreground="yellow")
            self.win.update()

            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            # Map ID -> first row
            id_row_map = {}
            for row in range(1, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=1).value
                if cell_val is not None:
                    key = str(cell_val).strip()
                    if key not in id_row_map:
                        id_row_map[key] = row

            success_count = 0
            not_found = []
            full_column = []
            col_start, col_end = 15, 61  # O=15, BI=61

            for search_id, detail_text in pairs:
                target_row = id_row_map.get(search_id)
                if target_row:
                    filled = False
                    for col_idx in range(col_start, col_end + 1):
                        cell = ws.cell(row=target_row, column=col_idx)
                        if cell.value is None or str(cell.value).strip() == "":
                            cell.value = detail_text
                            filled = True
                            success_count += 1
                            break
                    if not filled:
                        full_column.append(search_id)
                else:
                    not_found.append(search_id)

            base, ext = os.path.splitext(self.file_path)
            new_path = f"{base}_updated{ext}"
            wb.save(new_path)

            msg = f"เสร็จสิ้น! บันทึกไฟล์ที่: {os.path.basename(new_path)}\n\n"
            msg += f"สำเร็จ: {success_count} รายการ\n"
            if not_found:
                msg += f"ไม่พบรหัส: {len(not_found)} รายการ (เช่น {not_found[:3]})\n"
            if full_column:
                msg += f"ช่องเต็ม (O-BI): {len(full_column)} รายการ"

            self.lbl_status.configure(
                text=f"เสร็จสิ้น! (สำเร็จ {success_count} รายการ)", foreground="#2ECC71"
            )
            Messagebox.show_info(msg, title="Report")

            self.file_path = new_path
            self.entry_path.delete(0, "end")
            self.entry_path.insert(0, new_path)

        except Exception as e:
            Messagebox.show_error(f"เกิดข้อผิดพลาด: {str(e)}", title="Error")
            self.lbl_status.configure(text="เกิดข้อผิดพลาด", foreground="red")


# ============================================================================
# TOOL 2: กรอกข้อมูล Output Combined
# ============================================================================
class FillOutputWindow:
    """บันทึกรายงานการโทร -> Column G แล้วรวม F-U -> F"""

    def __init__(self, parent):
        self.parent = parent
        self.win = ttk.Toplevel(parent)
        self.win.title("กรอกข้อมูล Output Combined")
        self.status_var = ttk.StringVar(value="Ready")
        self._build_ui()
        fit_window(self.win, min_w=920, min_h=680)
        apply_bg(self.win)

    def _build_ui(self):
        main_frame = ttk.Frame(self.win)
        main_frame.pack(fill=BOTH, expand=YES, padx=SP, pady=SP)

        # Header
        header = ttk.Frame(main_frame, style="Card.TFrame")
        header.pack(fill=X, pady=(0, SP))
        ttk.Label(
            header,
            text="บันทึกรายงานการโทร + รวมคอลัมน์ F-U",
            font=(FONT_FAMILY, FONT_SIZE_TITLE, "bold"),
        ).pack(pady=SP)

        # Input area
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=BOTH, expand=YES)

        left_frame = ttk.LabelFrame(input_frame, text="รหัสลูกค้า", padding=SP)
        left_frame.pack(side=LEFT, fill=BOTH, expand=YES, padx=(0, SP // 2))
        self.text_id = ScrolledText(left_frame, width=25, height=24, autohide=True)
        self.text_id.pack(fill=BOTH, expand=YES)

        right_frame = ttk.LabelFrame(input_frame, text="รายงานการโทร", padding=SP)
        right_frame.pack(side=LEFT, fill=BOTH, expand=YES, padx=(SP // 2, 0))
        self.text_report = ScrolledText(
            right_frame, width=50, height=24, autohide=True
        )
        self.text_report.pack(fill=BOTH, expand=YES)

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=X, pady=SP)
        ttk.Button(
            btn_frame,
            text="Next >",
            command=self._next_page,
            bootstyle=PRIMARY,
        ).pack(fill=X, ipady=5)

        ttk.Label(main_frame, textvariable=self.status_var, foreground="gray").pack()

        # Keyboard shortcut
        self.win.bind("<Control-Return>", lambda e: self._next_page())

    def _next_page(self):
        customer_ids = self.text_id.get("1.0", "end").strip()
        call_reports = self.text_report.get("1.0", "end").strip()

        if not customer_ids or not call_reports:
            Messagebox.show_error("กรุณากรอกข้อมูลให้ครบถ้วน", title="Error")
            return

        cid_list = [c.strip() for c in customer_ids.splitlines() if c.strip()]
        report_list = [r.strip() for r in call_reports.splitlines() if r.strip()]

        if len(cid_list) != len(report_list):
            Messagebox.show_error(
                "จำนวนบรรทัดรหัสลูกค้าและรายงานการโทรต้องเท่ากัน", title="Error"
            )
            return

        self._show_table(cid_list, report_list)

    def _show_table(self, cid_list, report_list):
        table_win = ttk.Toplevel(self.win)
        table_win.title("ตรวจสอบข้อมูล")

        header_frame = ttk.Frame(table_win, style="Card.TFrame")
        header_frame.pack(fill=X, padx=SP, pady=SP)
        ttk.Label(
            header_frame,
            text="ตรวจสอบความถูกต้องก่อนบันทึก",
            font=(FONT_FAMILY, FONT_SIZE_TITLE, "bold"),
        ).pack(pady=SP)

        table_frame = ttk.Frame(table_win)
        table_frame.pack(fill=BOTH, expand=YES, padx=SP, pady=(0, SP))

        columns = ("customer_id", "call_report")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        tree.heading("customer_id", text="รหัสลูกค้า")
        tree.heading("call_report", text="รายงานการโทร")
        tree.column("customer_id", width=150, anchor="center")
        tree.column("call_report", width=550, anchor="w")

        for i, (cid, report) in enumerate(zip(cid_list, report_list)):
            truncated = (report[:50] + "...") if len(report) > 50 else report
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            tree.insert("", END, values=(cid, truncated), tags=(tag,))

        apply_table_style(tree)

        vsb = ttk.Scrollbar(table_frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        ttk.Button(
            table_win,
            text="ยืนยันและเลือกไฟล์ Excel",
            bootstyle=SUCCESS,
            command=lambda: (
                table_win.destroy(),
                self._select_excel(cid_list, report_list),
            ),
        ).pack(pady=SP)

        fit_window(table_win, min_w=760, min_h=680)
        apply_bg(table_win)

    def _select_excel(self, cid_list, report_list):
        file_path = filedialog.askopenfilename(
            title="เลือกไฟล์ Excel", filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return

        self.status_var.set("กำลังประมวลผล...")

        def process():
            try:
                output_path, not_found = self._process_update_and_combine(
                    file_path, cid_list, report_list
                )
                msg = f"บันทึกไฟล์เสร็จสิ้น: {os.path.basename(output_path)}"
                if not_found:
                    msg += f"\n\nไม่พบรหัสลูกค้า: {', '.join(not_found[:5])}"
                    if len(not_found) > 5:
                        msg += f" และอื่นๆ อีก {len(not_found) - 5} รายการ"

                show_toast("สำเร็จ", msg)
                self.status_var.set("เรียบร้อย: อัปเดตและรวมคอลัมน์แล้ว")
                self.win.destroy()
            except Exception as e:
                Messagebox.show_error(f"เกิดข้อผิดพลาด: {e}", title="Error")
                self.status_var.set("Ready")

        threading.Thread(target=process, daemon=True).start()

    @staticmethod
    def _process_update_and_combine(file_path, cid_list, report_list):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        not_found = []
        id_report_dict = dict(zip(cid_list, report_list))
        found_set = set()

        for row in ws.iter_rows(
            min_row=ROW_START, max_row=ROW_END, max_col=1, values_only=False
        ):
            cell = row[0]
            cid = str(cell.value).strip() if cell.value is not None else ""
            if cid in id_report_dict:
                ws.cell(row=cell.row, column=7, value=id_report_dict[cid])
                found_set.add(cid)

        for cid in cid_list:
            if cid not in found_set:
                not_found.append(cid)

        # Combine F(6)-U(21) -> F(6), clear G(7)-U(21)
        for row in range(ROW_START, ROW_END + 1):
            contents = []
            for col in range(6, 22):
                val = ws.cell(row=row, column=col).value
                if val is not None and str(val).strip() != "":
                    contents.append(str(val))
            ws.cell(row=row, column=6).value = "\n".join(contents)
            for col in range(7, 22):
                ws.cell(row=row, column=col).value = None

        output_path = os.path.splitext(file_path)[0] + "_processed.xlsx"
        wb.save(output_path)
        return output_path, not_found


# ============================================================================
# TOOL 3: ดึงข้อมูล 3T
# ============================================================================
class Extract3TWindow:
    """นับสายการโทร + เรียงลำดับ (Count / Date / ID)"""

    def __init__(self, parent):
        self.win = ttk.Toplevel(parent)
        self.win.title("ดึงข้อมูล 3T - นับสายการโทร")
        self.df_processed = None
        self._build_ui()
        fit_window(self.win, min_w=1100, min_h=620)
        apply_bg(self.win)

    def _build_ui(self):
        # --- Controls ---
        ctrl = ttk.Frame(self.win, style="Card.TFrame", padding=SP)
        ctrl.pack(fill=X, padx=SP, pady=SP)

        ttk.Button(
            ctrl, text="อัพโหลดไฟล์ Excel", command=self._load_excel, bootstyle=PRIMARY
        ).pack(side=LEFT, padx=5)

        self.btn_sort_count = ttk.Button(
            ctrl,
            text="เรียงตามจำนวน (น้อย->มาก)",
            command=lambda: self._sort_data("count"),
            state=DISABLED,
            bootstyle=SECONDARY,
        )
        self.btn_sort_count.pack(side=LEFT, padx=5)

        self.btn_sort_date = ttk.Button(
            ctrl,
            text="เรียงตามวันที่ (เก่า->ใหม่)",
            command=lambda: self._sort_data("date"),
            state=DISABLED,
            bootstyle=WARNING,
        )
        self.btn_sort_date.pack(side=LEFT, padx=5)

        self.btn_sort_id = ttk.Button(
            ctrl,
            text="เรียงตามรหัส",
            command=lambda: self._sort_data("id"),
            state=DISABLED,
            bootstyle=SECONDARY,
        )
        self.btn_sort_id.pack(side=LEFT, padx=5)

        self.btn_copy = ttk.Button(
            ctrl,
            text="Copy",
            command=self._copy_selection,
            state=DISABLED,
            bootstyle=INFO,
        )
        self.btn_copy.pack(side=LEFT, padx=20)

        self.lbl_status = ttk.Label(ctrl, text="สถานะ: รอไฟล์...", foreground="gray")
        self.lbl_status.pack(side=LEFT, padx=20)

        # --- Table ---
        table_frame = ttk.Frame(self.win)
        table_frame.pack(expand=YES, fill=BOTH, padx=SP, pady=(0, SP))

        columns = ("ID", "ColK", "Count", "Latest", "Details")
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", height=20, selectmode="extended"
        )

        self.tree.heading("ID", text="รหัสลูกค้า")
        self.tree.heading("ColK", text="ข้อมูล (Col K)")
        self.tree.heading("Count", text="จำนวน")
        self.tree.heading("Latest", text="ข้อความล่าสุด (วันที่+ผล)")
        self.tree.heading("Details", text="ข้อมูลอื่นๆ")

        self.tree.column("ID", width=100, anchor="center")
        self.tree.column("ColK", width=200, anchor="w")
        self.tree.column("Count", width=80, anchor="center")
        self.tree.column("Latest", width=250, anchor="w")
        self.tree.column("Details", width=300, anchor="w")

        vsb = ttk.Scrollbar(table_frame, orient=VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient=HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self.tree.bind("<<TreeviewSelect>>", self._on_selection_change)
        self.tree.bind("<Control-c>", lambda e: self._copy_selection())

        # Context menu
        self.context_menu = tk.Menu(self.win, tearoff=0)
        self.context_menu.add_command(label="Copy", command=self._copy_selection)
        self.context_menu.add_command(label="Copy All", command=self._copy_all)
        self.tree.bind("<Button-3>", self._show_context_menu)

    def _load_excel(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        try:
            self.lbl_status.config(text="กำลังประมวลผล...", foreground="blue")
            self.win.update()

            df = pd.read_excel(file_path, header=0)
            df_unique = df.drop_duplicates(subset=df.columns[0], keep="first").copy()

            # Call data range: O(14) - BI(60)
            call_data_range = df_unique.iloc[:, 14:61]
            count_vals = call_data_range.notna().sum(axis=1)

            def get_last_raw_value(row):
                last_col = row.last_valid_index()
                return str(row[last_col]) if last_col is not None else ""

            raw_latest = call_data_range.apply(get_last_raw_value, axis=1)

            def extract_date_for_sort(text_val):
                if not text_val:
                    return pd.NaT
                try:
                    return pd.to_datetime(
                        text_val.split()[0], dayfirst=True, errors="coerce"
                    )
                except Exception:
                    return pd.NaT

            sort_dates = raw_latest.apply(extract_date_for_sort)

            self.df_processed = pd.DataFrame(
                {
                    "customer_id": df_unique.iloc[:, 0],
                    "col_k_info": df_unique.iloc[:, 10].fillna(""),
                    "call_count": count_vals,
                    "latest_text": raw_latest,
                    "sort_date": sort_dates,
                    "other_info": df_unique.iloc[:, 1].fillna(""),
                }
            )

            self._update_table(self.df_processed)

            self.btn_sort_count.config(state=NORMAL)
            self.btn_sort_id.config(state=NORMAL)
            self.btn_sort_date.config(state=NORMAL)
            self.btn_copy.config(state=NORMAL)
            self.lbl_status.config(
                text=f"โหลดเสร็จ: {len(self.df_processed)} รายการ", foreground="green"
            )

        except Exception as e:
            Messagebox.show_error(f"เกิดข้อผิดพลาด: {e}", title="Error")

    def _sort_data(self, criteria):
        if self.df_processed is None:
            return

        df = self.df_processed.copy()

        if criteria == "count":
            df = df.sort_values(
                by=["call_count", "sort_date"], ascending=[True, True]
            )
            self.lbl_status.config(text="เรียง: จำนวนน้อย -> มาก")
        elif criteria == "id":
            df = df.sort_values(by="customer_id", ascending=True)
            self.lbl_status.config(text="เรียง: รหัสลูกค้า")
        elif criteria == "date":
            df = df.sort_values(by="sort_date", ascending=True, na_position="first")
            self.lbl_status.config(text="เรียง: วันที่เก่าสุด -> ล่าสุด")

        self._update_table(df)

    def _update_table(self, df):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for _, row in df.iterrows():
            self.tree.insert(
                "",
                "end",
                values=(
                    row["customer_id"],
                    row["col_k_info"],
                    row["call_count"],
                    row["latest_text"],
                    row["other_info"],
                ),
            )

    def _on_selection_change(self, event):
        cnt = len(self.tree.selection())
        if cnt > 0:
            self.lbl_status.config(text=f"เลือก: {cnt} รายการ")

    def _show_context_menu(self, event):
        self.context_menu.tk_popup(event.x_root, event.y_root)

    def _copy_selection(self):
        selected = self.tree.selection()
        if not selected:
            return
        text = ""
        for item in selected:
            vals = self.tree.item(item, "values")
            text += "\t".join(map(str, vals)) + "\n"
        self.win.clipboard_clear()
        self.win.clipboard_append(text)
        self.lbl_status.config(text="คัดลอกแล้ว!")

    def _copy_all(self):
        if not self.tree.get_children():
            return
        text = ""
        for item in self.tree.get_children():
            vals = self.tree.item(item, "values")
            text += "\t".join(map(str, vals)) + "\n"
        self.win.clipboard_clear()
        self.win.clipboard_append(text)
        self.lbl_status.config(text="คัดลอกทั้งหมดแล้ว!")


# ============================================================================
# TOOL 4: ดึงข้อมูลจาก Output Combined
# ============================================================================
class ExtractOutputWindow:
    """ค้นหารหัส + ดึงข้อมูล Column A, B, C, D, F"""

    def __init__(self, parent):
        self.win = ttk.Toplevel(parent)
        self.win.title("ดึงข้อมูลจาก Output Combined")
        self.file_path = None
        self._build_ui()
        fit_window(self.win, min_w=720, min_h=600)
        apply_bg(self.win)

    def _build_ui(self):
        # --- File selection ---
        file_frame = ttk.LabelFrame(self.win, text="1. เลือกไฟล์ Excel", padding=SP)
        file_frame.pack(fill=X, padx=SP, pady=SP)

        ttk.Button(
            file_frame, text="เลือกไฟล์", command=self._browse_file, bootstyle=PRIMARY
        ).pack(side=LEFT)

        self.lbl_file = ttk.Label(file_frame, text="...", foreground="gray")
        self.lbl_file.pack(side=LEFT, padx=SP)

        # --- Input ---
        input_frame = ttk.LabelFrame(
            self.win, text="2. กรอกรหัส (บรรทัดละ 1 รหัส)", padding=SP
        )
        input_frame.pack(fill=BOTH, expand=YES, padx=SP, pady=(0, SP))

        self.txt_input = ScrolledText(input_frame, height=12, autohide=True)
        self.txt_input.pack(fill=BOTH, expand=YES)

        # --- Process ---
        ttk.Button(
            self.win,
            text="ดึงข้อมูล",
            command=self._process_data,
            bootstyle=SUCCESS,
        ).pack(fill=X, padx=SP, pady=(0, SP), ipady=5)

        self.lbl_status = ttk.Label(self.win, text="พร้อมทำงาน", foreground="gray")
        self.lbl_status.pack(pady=(0, SP))

    def _browse_file(self):
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.file_path = filename
            self.lbl_file.config(text=os.path.basename(filename), foreground="white")

    def _process_data(self):
        if not self.file_path:
            Messagebox.show_warning("กรุณาเลือกไฟล์ Excel ก่อน", title="เตือน")
            return

        raw_text = self.txt_input.get("1.0", "end")
        search_codes = [line.strip() for line in raw_text.splitlines() if line.strip()]

        if not search_codes:
            Messagebox.show_warning("กรุณากรอกรหัสอย่างน้อย 1 ตัว", title="เตือน")
            return

        try:
            df = pd.read_excel(self.file_path, header=0)
            df.iloc[:, 0] = df.iloc[:, 0].astype(str)

            found_rows = []
            for code in search_codes:
                matched = df[df.iloc[:, 0] == code]
                if not matched.empty:
                    if len(matched.columns) > 5:
                        found_rows.append(matched.iloc[:, [0, 1, 2, 3, 5]])
                    else:
                        found_rows.append(matched)

            if not found_rows:
                Messagebox.show_info("ไม่พบข้อมูลรหัสที่ระบุ", title="ผลลัพธ์")
                return

            final_df = pd.concat(found_rows).fillna("")
            result_text = final_df.to_csv(
                sep="\t", index=False, header=False, lineterminator="\n"
            ).strip()

            self._show_result(result_text)

        except Exception as e:
            Messagebox.show_error(f"เกิดข้อผิดพลาด: {str(e)}", title="Error")

    def _show_result(self, text_data):
        result_win = ttk.Toplevel(self.win)
        result_win.title("ผลลัพธ์ (พร้อม Copy)")

        ttk.Label(
            result_win,
            text="กดปุ่ม Copy แล้วไปวางใน Excel:",
            font=(FONT_FAMILY, FONT_SIZE, "bold"),
        ).pack(pady=SP)

        txt = ScrolledText(result_win, height=20, autohide=True)
        txt.pack(fill=BOTH, expand=YES, padx=SP, pady=(0, SP))
        txt.insert("end", text_data)

        def copy_all():
            result_win.clipboard_clear()
            result_win.clipboard_append(text_data)
            result_win.update()
            Messagebox.show_info(
                "คัดลอกแล้ว!\nไปที่ Excel แล้วกด Paste (Ctrl+V) ได้เลย", title="OK"
            )

        ttk.Button(
            result_win, text="Copy All", command=copy_all, bootstyle=WARNING
        ).pack(fill=X, padx=SP * 2, pady=SP, ipady=5)

        fit_window(result_win, min_w=720, min_h=480)
        apply_bg(result_win)


# ============================================================================
# TOOL 5: เปิดบิล (Bill Workflow)
# ============================================================================
def _bill_format_values(values, num_columns, paper_width):
    """Format proof paper roll weights into a column-grid layout."""
    total_count = len(values)
    total_sum = sum(values)
    num_rows = (total_count + num_columns - 1) // num_columns
    grid = [[] for _ in range(num_rows)]
    for i, value in enumerate(values):
        grid[i % num_rows].append(f"{i + 1} - {value:.1f} kg")
    header = f"กระดาษรองรีดหน้า {paper_width} นิ้ว หนา 48.8 แกรม"
    rows = [" ".join(row) for row in grid]
    total = f"รวมทั้งหมด {total_count} ม้วน = {total_sum:.1f} kg"
    return "\n".join([header] + rows + [total])


class SublimationWindow:
    """กระดาษซับ Sublimation Paper — คำนวณราคาและสร้างบิล"""

    def __init__(self, parent):
        self.win = ttk.Toplevel(parent)
        self.win.title("กระดาษซับ (Sublimation Paper)")
        self._build_ui()
        fit_window(self.win, min_w=700, min_h=640)
        apply_bg(self.win)

    def _build_ui(self):
        ttk.Label(
            self.win, text="กระดาษซับ (Sublimation Paper)",
            font=(FONT_FAMILY, 15, "bold"),
        ).pack(pady=(SP, 0))

        form = ttk.LabelFrame(self.win, text="ข้อมูลกระดาษ", padding=SP)
        form.pack(fill=X, padx=SP * 2, pady=SP)

        fields = [
            ("ความกว้าง (ซม.):", "ent_width"),
            ("ความหนา (แกรม):", "ent_thick"),
            ("ความยาว (เมตร):", "ent_len"),
            ("ราคาต่อม้วน (บาท):", "ent_price"),
            ("จำนวนม้วน:", "ent_rolls"),
            ("จำนวน Column:", "ent_cols"),
        ]
        for i, (label, attr) in enumerate(fields):
            ttk.Label(form, text=label, width=25, anchor=W).grid(
                row=i, column=0, sticky=W, pady=4, padx=5
            )
            ent = ttk.Entry(form, width=22)
            ent.grid(row=i, column=1, pady=4, padx=5, sticky=W)
            setattr(self, attr, ent)

        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=SP)
        ttk.Button(btn_frame, text="Submit", command=self._submit,
                   bootstyle=SUCCESS).pack(side=LEFT, padx=5)
        self.copy_btn = ttk.Button(btn_frame, text="Copy",
                                    command=self._copy, bootstyle=INFO)
        self.copy_btn.pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="Clear", command=self._clear,
                   bootstyle=WARNING).pack(side=LEFT, padx=5)

        out = ttk.LabelFrame(self.win, text="ผลลัพธ์", padding=SP)
        out.pack(fill=BOTH, expand=YES, padx=SP * 2, pady=(0, SP))
        self.output = ScrolledText(out, height=14, autohide=True)
        self.output.pack(fill=BOTH, expand=YES)

    def _submit(self):
        try:
            width = self.ent_width.get().strip()
            thick = self.ent_thick.get().strip()
            length = self.ent_len.get().strip()
            price = float(self.ent_price.get().strip())
            num_rolls = int(self.ent_rolls.get().strip())
            num_cols = int(self.ent_cols.get().strip())

            header = f"กระดาษซับ {width} ซ.ม. {thick} แกรม {length} เมตร"
            num_rows = (num_rolls + num_cols - 1) // num_cols
            grid = [[] for _ in range(num_rows)]
            for i in range(num_rolls):
                grid[i % num_rows].append(f"{i + 1} - {price:,.1f} บาท")

            total = num_rolls * price
            summary = f"รวมทั้งหมด {num_rolls:,} ม้วน = {total:,.1f} บาท"
            lines = [header] + [" ".join(r) for r in grid] + [summary]

            self.output.delete("1.0", "end")
            self.output.insert("end", "\n".join(lines))
        except ValueError:
            Messagebox.show_error("ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบอีกครั้ง", title="Error")

    def _copy(self):
        text = self.output.get("1.0", "end").strip()
        self.win.clipboard_clear()
        self.win.clipboard_append(text)
        self.copy_btn.config(text="Copied!")
        self.win.after(1000, lambda: self.copy_btn.config(text="Copy"))

    def _clear(self):
        for attr in ("ent_width", "ent_thick", "ent_len", "ent_price",
                     "ent_rolls", "ent_cols"):
            getattr(self, attr).delete(0, "end")
        self.output.delete("1.0", "end")


class ProofWindow:
    """กระดาษรองรีด Proof Paper — กรอกน้ำหนักม้วนสร้างบิล"""

    def __init__(self, parent):
        self.win = ttk.Toplevel(parent)
        self.win.title("กระดาษรองรีด (Proof Paper)")
        self._build_ui()
        fit_window(self.win, min_w=700, min_h=660)
        apply_bg(self.win)

    def _build_ui(self):
        ttk.Label(
            self.win, text="กระดาษรองรีด (Proof Paper)",
            font=(FONT_FAMILY, 15, "bold"),
        ).pack(pady=(SP, 0))

        form = ttk.LabelFrame(self.win, text="ข้อมูลกระดาษ", padding=SP)
        form.pack(fill=X, padx=SP * 2, pady=SP)

        # Row: width
        r0 = ttk.Frame(form)
        r0.pack(fill=X, pady=4)
        ttk.Label(r0, text="ความกว้าง (นิ้ว):", width=30, anchor=W).pack(side=LEFT)
        self.ent_width = ttk.Entry(r0, width=22)
        self.ent_width.pack(side=LEFT, padx=5)

        # Row: weights textarea
        r1 = ttk.Frame(form)
        r1.pack(fill=X, pady=4)
        ttk.Label(r1, text="น้ำหนักม้วน (kg) บรรทัดละ 1 ค่า:", width=30, anchor=NW).pack(side=LEFT)
        self.txt_weights = ScrolledText(r1, height=8, width=20, autohide=True)
        self.txt_weights.pack(side=LEFT, padx=5)

        # Row: columns
        r2 = ttk.Frame(form)
        r2.pack(fill=X, pady=4)
        ttk.Label(r2, text="จำนวน Column:", width=30, anchor=W).pack(side=LEFT)
        self.ent_cols = ttk.Entry(r2, width=22)
        self.ent_cols.pack(side=LEFT, padx=5)

        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=SP)
        ttk.Button(btn_frame, text="Submit", command=self._submit,
                   bootstyle=SUCCESS).pack(side=LEFT, padx=5)
        self.copy_btn = ttk.Button(btn_frame, text="Copy",
                                    command=self._copy, bootstyle=INFO)
        self.copy_btn.pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="Clear", command=self._clear,
                   bootstyle=WARNING).pack(side=LEFT, padx=5)

        out = ttk.LabelFrame(self.win, text="ผลลัพธ์", padding=SP)
        out.pack(fill=BOTH, expand=YES, padx=SP * 2, pady=(0, SP))
        self.output = ScrolledText(out, height=10, autohide=True)
        self.output.pack(fill=BOTH, expand=YES)

    def _submit(self):
        try:
            width = self.ent_width.get().strip()
            weights_raw = self.txt_weights.get("1.0", "end").strip()
            num_cols = int(self.ent_cols.get().strip())
            weights = [float(w) for w in weights_raw.splitlines() if w.strip()]
            result = _bill_format_values(weights, num_cols, width)
            self.output.delete("1.0", "end")
            self.output.insert("end", result)
        except ValueError:
            Messagebox.show_error("ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบอีกครั้ง", title="Error")

    def _copy(self):
        text = self.output.get("1.0", "end").strip()
        self.win.clipboard_clear()
        self.win.clipboard_append(text)
        self.copy_btn.config(text="Copied!")
        self.win.after(1000, lambda: self.copy_btn.config(text="Copy"))

    def _clear(self):
        self.ent_width.delete(0, "end")
        self.txt_weights.delete("1.0", "end")
        self.ent_cols.delete(0, "end")
        self.output.delete("1.0", "end")


# ============================================================================
# MAIN APPLICATION (Home Screen)
# ============================================================================
class MainApp(ttk.Window):
    def __init__(self):
        super().__init__(themename="darkly")
        self.title("Excel All-in-One Tool")
        center_window(self, 860, 820)
        self.option_add("*Font", (FONT_FAMILY, FONT_SIZE))
        self._btn_color_map = []   # [(widget, base_bg), ...]
        apply_bg(self)
        self._build_ui()
        # บังคับสีหลัง ttkbootstrap theme render เสร็จ
        self.after(100, self._enforce_btn_colors)

    def _enforce_btn_colors(self):
        for widget, color in self._btn_color_map:
            try:
                widget.configure(bg=color)
            except Exception:
                pass

    def _build_ui(self):
        # --- Header ---
        ttk.Label(
            self, text="Excel All-in-One Tool", font=(FONT_FAMILY, 22, "bold")
        ).pack(pady=(SP + 4, 0))
        ttk.Label(
            self, text="เลือกเครื่องมือที่ต้องการใช้งาน",
            foreground="gray", font=(FONT_FAMILY, FONT_SIZE),
        ).pack(pady=(4, SP))

        # --- กรอกข้อมูล ---
        grp_fill = ttk.LabelFrame(self, text="  กรอกข้อมูล  ", padding=(SP, SP // 2))
        grp_fill.pack(fill=X, padx=SP * 2, pady=(0, SP))

        self._custom_btn(
            grp_fill,
            prefix="กรอกข้อมูล",
            keyword="3T",
            kw_color="#1976D2",
            suffix="กรอกรหัส + รายละเอียด ลง Column O-BI",
            command=lambda: Fill3TWindow(self),
            base_bg="#1a3a6e",
            hover_bg="#1f4a8a",
        )
        self._custom_btn(
            grp_fill,
            prefix="กรอกข้อมูล",
            keyword="Output",
            kw_color="#C62828",
            suffix="บันทึกรายงานการโทร + รวม Column F-U -> F",
            command=lambda: FillOutputWindow(self),
            base_bg="#6e1a1a",
            hover_bg="#8a2020",
        )

        # --- ดึงข้อมูล ---
        grp_extract = ttk.LabelFrame(self, text="  ดึงข้อมูล  ", padding=(SP, SP // 2))
        grp_extract.pack(fill=X, padx=SP * 2, pady=(0, SP))

        self._custom_btn(
            grp_extract,
            prefix="ดึงข้อมูล",
            keyword="3T",
            kw_color="#1976D2",
            suffix="นับสายการโทร + เรียงลำดับ (Count / Date / ID)",
            command=lambda: Extract3TWindow(self),
            base_bg="#1a3a6e",
            hover_bg="#1f4a8a",
        )
        self._custom_btn(
            grp_extract,
            prefix="ดึงข้อมูล",
            keyword="Output",
            kw_color="#C62828",
            suffix="ค้นหารหัส + ดึงข้อมูล Column A, B, C, D, F",
            command=lambda: ExtractOutputWindow(self),
            base_bg="#6e1a1a",
            hover_bg="#8a2020",
        )

        # --- เปิดบิล ---
        grp_bill = ttk.LabelFrame(self, text="  เปิดบิล  ", padding=(SP, SP // 2))
        grp_bill.pack(fill=X, padx=SP * 2, pady=(0, SP))

        self._custom_btn(
            grp_bill,
            prefix="กระดาษ",
            keyword="ซับ",
            kw_color="#6A1B9A",
            suffix="กรอกขนาด/ราคา/จำนวนม้วน → สร้างบิล",
            command=lambda: SublimationWindow(self),
            base_bg="#3a1060",
            hover_bg="#4a1580",
        )
        self._custom_btn(
            grp_bill,
            prefix="กระดาษ",
            keyword="รองรีด",
            kw_color="#7B1FA2",
            suffix="กรอกน้ำหนักม้วน (kg) → สร้างบิล",
            command=lambda: ProofWindow(self),
            base_bg="#3a1060",
            hover_bg="#4a1580",
        )
        # --- Footer ---
        ttk.Label(
            self, text="v5.0 All-in-One", foreground="gray", font=(FONT_FAMILY, 9)
        ).pack(side=BOTTOM, pady=SP)

    def _custom_btn(self, parent, prefix, keyword, kw_color, suffix, command, base_bg, hover_bg):
        """Custom button: left accent line + keyword badge + description text"""
        outer = tk.Frame(parent, bg=base_bg, cursor="hand2")
        outer.pack(fill=X, pady=(0, 6))

        # Left color accent bar
        accent = tk.Frame(outer, bg=kw_color, width=5)
        accent.pack(side=LEFT, fill=Y)

        inner = tk.Frame(outer, bg=base_bg)
        inner.pack(side=LEFT, fill=X, expand=YES, padx=14, pady=11)

        pre_lbl = tk.Label(
            inner, text=prefix + "  ",
            bg=base_bg, fg="#eeeeee",
            font=(FONT_FAMILY, 12),
        )
        pre_lbl.pack(side=LEFT)

        badge = tk.Label(
            inner, text=f"  {keyword}  ",
            bg=kw_color, fg="white",
            font=(FONT_FAMILY, 13, "bold"),
            pady=1,
        )
        badge.pack(side=LEFT)

        sep_lbl = tk.Label(
            inner, text="   —   ",
            bg=base_bg, fg="#777777",
            font=(FONT_FAMILY, 11),
        )
        sep_lbl.pack(side=LEFT)

        desc_lbl = tk.Label(
            inner, text=suffix,
            bg=base_bg, fg="#bbbbbb",
            font=(FONT_FAMILY, 11),
        )
        desc_lbl.pack(side=LEFT)

        all_bg = [outer, inner, pre_lbl, sep_lbl, desc_lbl]

        # Register for color enforcement on startup
        for w in all_bg:
            self._btn_color_map.append((w, base_bg))

        def on_enter(e):
            for w in all_bg:
                w.config(bg=hover_bg)

        def on_leave(e):
            # Only reset when mouse truly leaves the outer frame
            rx, ry = e.x_root, e.y_root
            ox = outer.winfo_rootx()
            oy = outer.winfo_rooty()
            if not (ox <= rx < ox + outer.winfo_width() and
                    oy <= ry < oy + outer.winfo_height()):
                for w in all_bg:
                    w.config(bg=base_bg)

        def on_click(e):
            command()

        for w in [*all_bg, accent, badge]:
            w.bind("<Enter>", on_enter)
            w.bind("<Leave>", on_leave)
            w.bind("<Button-1>", on_click)


# ============================================================================
# ENTRY POINT
# ============================================================================
if __name__ == "__main__":
    app = MainApp()
    app.mainloop()
