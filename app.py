# -*- coding: utf-8 -*-
"""
Excel All-in-One Tool v5.0 — Streamlit Edition
รวม 5 เครื่องมือ:
1. กรอกข้อมูล 3T
2. กรอกข้อมูล Output
3. ดึงข้อมูล 3T
4. ดึงข้อมูลจาก Output
5. เปิดบิล (กระดาษซับ / กระดาษรองรีด)
"""

import io
import streamlit as st
import openpyxl
import pandas as pd

st.set_page_config(page_title="Excel All-in-One Tool", page_icon="📊", layout="wide")

# ============================================================
# SIDEBAR — เมนูหลัก
# ============================================================
st.sidebar.title("📊 Excel All-in-One Tool")
st.sidebar.markdown("---")
tool = st.sidebar.radio(
    "เลือกเครื่องมือ",
    [
        "🏠 หน้าหลัก",
        "1️⃣  กรอกข้อมูล 3T",
        "2️⃣  กรอกข้อมูล Output",
        "3️⃣  ดึงข้อมูล 3T",
        "4️⃣  ดึงข้อมูลจาก Output",
        "5️⃣  เปิดบิล — กระดาษซับ",
        "6️⃣  เปิดบิล — กระดาษรองรีด",
    ],
)
st.sidebar.markdown("---")
st.sidebar.caption("v5.0 All-in-One · Streamlit Edition")

ROW_START = 2
ROW_END = 553


# ============================================================
# HELPER
# ============================================================
def to_excel_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# หน้าหลัก
# ============================================================
if tool == "🏠 หน้าหลัก":
    st.title("📊 Excel All-in-One Tool")
    st.markdown("### เลือกเครื่องมือจากเมนูด้านซ้าย")
    st.markdown("""
| # | เครื่องมือ | รายละเอียด |
|---|---|---|
| 1 | กรอกข้อมูล 3T | กรอกรหัส + รายละเอียด ลง Column O-BI |
| 2 | กรอกข้อมูล Output | บันทึกรายงานการโทร + รวม Column F-U → F |
| 3 | ดึงข้อมูล 3T | นับสายการโทร + เรียงลำดับ |
| 4 | ดึงข้อมูลจาก Output | ค้นหารหัส + ดึง Column A, B, C, D, F |
| 5 | เปิดบิล — กระดาษซับ | คำนวณราคา + สร้างบิล |
| 6 | เปิดบิล — กระดาษรองรีด | กรอกน้ำหนักม้วน + สร้างบิล |
""")


# ============================================================
# TOOL 1: กรอกข้อมูล 3T
# ============================================================
elif tool == "1️⃣  กรอกข้อมูล 3T":
    st.header("1️⃣ กรอกข้อมูล 3T")
    st.caption("กรอกรหัส + รายละเอียด ลง Column O-BI ของไฟล์ Excel")

    uploaded = st.file_uploader("อัปโหลดไฟล์ Excel (.xlsx / .xlsm)", type=["xlsx", "xlsm"])

    col1, col2 = st.columns(2)
    with col1:
        raw_ids = st.text_area("กรอก 'รหัส' (Column A) — บรรทัดละ 1 รหัส", height=300)
    with col2:
        raw_details = st.text_area("กรอก 'รายละเอียด' ที่จะเติม — บรรทัดละ 1 ค่า", height=300)

    if st.button("▶ เริ่มทำงาน", type="primary"):
        if not uploaded:
            st.error("กรุณาอัปโหลดไฟล์ Excel ก่อน")
        else:
            list_ids = [x.strip() for x in raw_ids.strip().splitlines() if x.strip()]
            list_details = [x.strip() for x in raw_details.strip().splitlines() if x.strip()]

            if not list_ids or not list_details:
                st.error("กรุณากรอกข้อมูลทั้งสองช่อง")
            else:
                if len(list_ids) != len(list_details):
                    st.warning(
                        f"จำนวนบรรทัดไม่เท่ากัน! รหัส: {len(list_ids)} / รายละเอียด: {len(list_details)}\n"
                        "โปรแกรมจะทำเฉพาะคู่ที่มีครบ"
                    )

                pairs = list(zip(list_ids, list_details))
                with st.spinner("กำลังประมวลผล..."):
                    wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()))
                    ws = wb.active

                    id_row_map = {}
                    for row in range(1, ws.max_row + 1):
                        cell_val = ws.cell(row=row, column=1).value
                        if cell_val is not None:
                            key = str(cell_val).strip()
                            if key not in id_row_map:
                                id_row_map[key] = row

                    success_count = 0
                    not_found = []
                    full_column = []
                    col_start, col_end = 15, 61  # O=15, BI=61

                    for search_id, detail_text in pairs:
                        target_row = id_row_map.get(search_id)
                        if target_row:
                            filled = False
                            for col_idx in range(col_start, col_end + 1):
                                cell = ws.cell(row=target_row, column=col_idx)
                                if cell.value is None or str(cell.value).strip() == "":
                                    cell.value = detail_text
                                    filled = True
                                    success_count += 1
                                    break
                            if not filled:
                                full_column.append(search_id)
                        else:
                            not_found.append(search_id)

                    result_bytes = to_excel_bytes(wb)

                st.success(f"เสร็จสิ้น! สำเร็จ {success_count} รายการ")
                if not_found:
                    st.warning(f"ไม่พบรหัส {len(not_found)} รายการ: {not_found[:5]}")
                if full_column:
                    st.warning(f"ช่องเต็ม (O-BI) {len(full_column)} รายการ: {full_column[:5]}")

                fname = uploaded.name.replace(".xlsx", "_updated.xlsx").replace(".xlsm", "_updated.xlsm")
                st.download_button("⬇ ดาวน์โหลดไฟล์ที่อัปเดตแล้ว", result_bytes, file_name=fname)


# ============================================================
# TOOL 2: กรอกข้อมูล Output
# ============================================================
elif tool == "2️⃣  กรอกข้อมูล Output":
    st.header("2️⃣ กรอกข้อมูล Output")
    st.caption("บันทึกรายงานการโทร → Column G แล้วรวม Column F-U → F")

    uploaded = st.file_uploader("อัปโหลดไฟล์ Excel (.xlsx)", type=["xlsx"])

    col1, col2 = st.columns(2)
    with col1:
        raw_ids = st.text_area("รหัสลูกค้า (บรรทัดละ 1 รหัส)", height=300)
    with col2:
        raw_reports = st.text_area("รายงานการโทร (บรรทัดละ 1 รายการ)", height=300)

    if st.button("▶ ประมวลผล", type="primary"):
        if not uploaded:
            st.error("กรุณาอัปโหลดไฟล์ Excel ก่อน")
        else:
            cid_list = [x.strip() for x in raw_ids.strip().splitlines() if x.strip()]
            report_list = [x.strip() for x in raw_reports.strip().splitlines() if x.strip()]

            if not cid_list or not report_list:
                st.error("กรุณากรอกข้อมูลทั้งสองช่อง")
            elif len(cid_list) != len(report_list):
                st.error(f"จำนวนบรรทัดต้องเท่ากัน! รหัส: {len(cid_list)} / รายงาน: {len(report_list)}")
            else:
                # Preview
                st.markdown("#### ตรวจสอบข้อมูลก่อนบันทึก")
                preview_df = pd.DataFrame({"รหัสลูกค้า": cid_list, "รายงานการโทร": report_list})
                st.dataframe(preview_df, use_container_width=True)

                with st.spinner("กำลังประมวลผล..."):
                    wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()))
                    ws = wb.active

                    id_report_dict = dict(zip(cid_list, report_list))
                    found_set = set()
                    not_found = []

                    for row in ws.iter_rows(min_row=ROW_START, max_row=ROW_END, max_col=1):
                        cell = row[0]
                        cid = str(cell.value).strip() if cell.value is not None else ""
                        if cid in id_report_dict:
                            ws.cell(row=cell.row, column=7, value=id_report_dict[cid])
                            found_set.add(cid)

                    for cid in cid_list:
                        if cid not in found_set:
                            not_found.append(cid)

                    # Combine F(6)-U(21) -> F(6)
                    for row in range(ROW_START, ROW_END + 1):
                        contents = []
                        for col in range(6, 22):
                            val = ws.cell(row=row, column=col).value
                            if val is not None and str(val).strip() != "":
                                contents.append(str(val))
                        ws.cell(row=row, column=6).value = "\n".join(contents)
                        for col in range(7, 22):
                            ws.cell(row=row, column=col).value = None

                    result_bytes = to_excel_bytes(wb)

                st.success("เสร็จสิ้น! รวมคอลัมน์ F-U → F แล้ว")
                if not_found:
                    st.warning(f"ไม่พบรหัสลูกค้า: {not_found}")

                fname = uploaded.name.replace(".xlsx", "_processed.xlsx")
                st.download_button("⬇ ดาวน์โหลดไฟล์ที่ประมวลผลแล้ว", result_bytes, file_name=fname)


# ============================================================
# TOOL 3: ดึงข้อมูล 3T
# ============================================================
elif tool == "3️⃣  ดึงข้อมูล 3T":
    st.header("3️⃣ ดึงข้อมูล 3T")
    st.caption("นับสายการโทร + เรียงลำดับ (Count / Date / ID)")

    uploaded = st.file_uploader("อัปโหลดไฟล์ Excel (.xlsx / .xls)", type=["xlsx", "xls"])

    if uploaded:
        with st.spinner("กำลังโหลดและประมวลผลไฟล์..."):
            df = pd.read_excel(uploaded, header=0)
            df_unique = df.drop_duplicates(subset=df.columns[0], keep="first").copy()

            call_data_range = df_unique.iloc[:, 14:61]
            count_vals = call_data_range.notna().sum(axis=1)

            def get_last_raw_value(row):
                last_col = row.last_valid_index()
                return str(row[last_col]) if last_col is not None else ""

            raw_latest = call_data_range.apply(get_last_raw_value, axis=1)

            def extract_date_for_sort(text_val):
                if not text_val:
                    return pd.NaT
                try:
                    return pd.to_datetime(text_val.split()[0], dayfirst=True, errors="coerce")
                except Exception:
                    return pd.NaT

            sort_dates = raw_latest.apply(extract_date_for_sort)

            result_df = pd.DataFrame({
                "รหัสลูกค้า": df_unique.iloc[:, 0],
                "ข้อมูล (Col K)": df_unique.iloc[:, 10].fillna(""),
                "จำนวน": count_vals,
                "ข้อความล่าสุด": raw_latest,
                "sort_date": sort_dates,
                "ข้อมูลอื่นๆ": df_unique.iloc[:, 1].fillna(""),
            })

        st.success(f"โหลดเสร็จ: {len(result_df)} รายการ")

        sort_by = st.selectbox(
            "เรียงลำดับตาม",
            ["ค่าเดิม", "จำนวน (น้อย→มาก)", "วันที่ (เก่า→ใหม่)", "รหัสลูกค้า"],
        )

        display_df = result_df.copy()
        if sort_by == "จำนวน (น้อย→มาก)":
            display_df = display_df.sort_values(by=["จำนวน", "sort_date"], ascending=[True, True])
        elif sort_by == "วันที่ (เก่า→ใหม่)":
            display_df = display_df.sort_values(by="sort_date", ascending=True, na_position="first")
        elif sort_by == "รหัสลูกค้า":
            display_df = display_df.sort_values(by="รหัสลูกค้า", ascending=True)

        show_df = display_df.drop(columns=["sort_date"])
        st.dataframe(show_df, use_container_width=True, height=500)

        # Download as Excel
        buf = io.BytesIO()
        show_df.to_excel(buf, index=False)
        st.download_button("⬇ ดาวน์โหลดผลลัพธ์ (.xlsx)", buf.getvalue(), file_name="3T_result.xlsx")


# ============================================================
# TOOL 4: ดึงข้อมูลจาก Output
# ============================================================
elif tool == "4️⃣  ดึงข้อมูลจาก Output":
    st.header("4️⃣ ดึงข้อมูลจาก Output")
    st.caption("ค้นหารหัส + ดึงข้อมูล Column A, B, C, D, F")

    uploaded = st.file_uploader("อัปโหลดไฟล์ Excel (.xlsx / .xls)", type=["xlsx", "xls"])
    raw_codes = st.text_area("กรอกรหัสที่ต้องการค้นหา (บรรทัดละ 1 รหัส)", height=250)

    if st.button("▶ ดึงข้อมูล", type="primary"):
        if not uploaded:
            st.error("กรุณาอัปโหลดไฟล์ Excel ก่อน")
        elif not raw_codes.strip():
            st.error("กรุณากรอกรหัสอย่างน้อย 1 รหัส")
        else:
            search_codes = [x.strip() for x in raw_codes.strip().splitlines() if x.strip()]
            with st.spinner("กำลังค้นหา..."):
                df = pd.read_excel(uploaded, header=0)
                df.iloc[:, 0] = df.iloc[:, 0].astype(str)

                found_rows = []
                for code in search_codes:
                    matched = df[df.iloc[:, 0] == code]
                    if not matched.empty:
                        if len(matched.columns) > 5:
                            found_rows.append(matched.iloc[:, [0, 1, 2, 3, 5]])
                        else:
                            found_rows.append(matched)

            if not found_rows:
                st.warning("ไม่พบข้อมูลรหัสที่ระบุ")
            else:
                final_df = pd.concat(found_rows).fillna("")
                final_df.columns = [f"Col{i+1}" for i in range(len(final_df.columns))]
                st.success(f"พบ {len(final_df)} แถว")
                st.dataframe(final_df, use_container_width=True)

                # TSV for pasting into Excel
                tsv = final_df.to_csv(sep="\t", index=False, header=False)
                st.text_area("📋 คัดลอกเพื่อวางใน Excel (Ctrl+A → Ctrl+C)", tsv, height=200)

                buf = io.BytesIO()
                final_df.to_excel(buf, index=False)
                st.download_button("⬇ ดาวน์โหลดผลลัพธ์ (.xlsx)", buf.getvalue(), file_name="output_result.xlsx")


# ============================================================
# TOOL 5: เปิดบิล — กระดาษซับ
# ============================================================
elif tool == "5️⃣  เปิดบิล — กระดาษซับ":
    st.header("5️⃣ กระดาษซับ (Sublimation Paper)")
    st.caption("คำนวณราคา + สร้างบิล")

    col1, col2 = st.columns(2)
    with col1:
        width = st.text_input("ความกว้าง (ซม.)")
        thick = st.text_input("ความหนา (แกรม)")
        length = st.text_input("ความยาว (เมตร)")
    with col2:
        price = st.number_input("ราคาต่อม้วน (บาท)", min_value=0.0, step=0.1)
        num_rolls = st.number_input("จำนวนม้วน", min_value=1, step=1, value=1)
        num_cols = st.number_input("จำนวน Column ในบิล", min_value=1, step=1, value=3)

    if st.button("▶ สร้างบิล", type="primary"):
        if not width or not thick or not length:
            st.error("กรุณากรอกข้อมูลให้ครบ")
        else:
            header_line = f"กระดาษซับ {width} ซ.ม. {thick} แกรม {length} เมตร"
            num_rows = (num_rolls + num_cols - 1) // num_cols
            grid = [[] for _ in range(num_rows)]
            for i in range(num_rolls):
                grid[i % num_rows].append(f"{i + 1} - {price:,.1f} บาท")
            total = num_rolls * price
            summary = f"รวมทั้งหมด {num_rolls:,} ม้วน = {total:,.1f} บาท"
            lines = [header_line] + ["   ".join(r) for r in grid] + [summary]
            result = "\n".join(lines)

            st.success("สร้างบิลเรียบร้อย!")
            st.text_area("📋 ผลลัพธ์บิล (คัดลอกได้เลย)", result, height=300)


# ============================================================
# TOOL 6: เปิดบิล — กระดาษรองรีด
# ============================================================
elif tool == "6️⃣  เปิดบิล — กระดาษรองรีด":
    st.header("6️⃣ กระดาษรองรีด (Proof Paper)")
    st.caption("กรอกน้ำหนักม้วน (kg) → สร้างบิล")

    paper_width = st.text_input("ความกว้าง (นิ้ว)")
    raw_weights = st.text_area("น้ำหนักม้วน (kg) — บรรทัดละ 1 ค่า", height=200)
    num_cols = st.number_input("จำนวน Column ในบิล", min_value=1, step=1, value=3)

    if st.button("▶ สร้างบิล", type="primary"):
        if not paper_width or not raw_weights.strip():
            st.error("กรุณากรอกข้อมูลให้ครบ")
        else:
            try:
                weights = [float(w) for w in raw_weights.strip().splitlines() if w.strip()]
                total_count = len(weights)
                total_sum = sum(weights)
                num_rows = (total_count + num_cols - 1) // num_cols
                grid = [[] for _ in range(num_rows)]
                for i, value in enumerate(weights):
                    grid[i % num_rows].append(f"{i + 1} - {value:.1f} kg")
                header_line = f"กระดาษรองรีดหน้า {paper_width} นิ้ว หนา 48.8 แกรม"
                summary = f"รวมทั้งหมด {total_count} ม้วน = {total_sum:.1f} kg"
                lines = [header_line] + ["   ".join(r) for r in grid] + [summary]
                result = "\n".join(lines)

                st.success("สร้างบิลเรียบร้อย!")
                st.text_area("📋 ผลลัพธ์บิล (คัดลอกได้เลย)", result, height=300)
            except ValueError:
                st.error("ข้อมูลน้ำหนักไม่ถูกต้อง กรุณากรอกตัวเลขเท่านั้น")
